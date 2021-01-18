# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Login.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
return_answer = 0
Dialog = QtWidgets.QWidget

# 암호화 및 복호화 세팅
from base64 import b64encode, b64decode
from Crypto.Cipher import AES

# 임의의 값으로 패딩
BS = 16
pad = lambda s: s + (BS - len(s) % BS) * chr(BS - len(s) % BS).encode()
unpad = lambda s: s[:-ord(s[len(s) - 1:])]


def iv():
    return chr(0) * 16


class AESCipher(object):
    def __init__(self, key):
        self.key = key

    # encrypt : 메시지를 암호화 하는 함수
    def encrypt(self, message):
        message = message.encode()
        raw = pad(message)
        cipher = AES.new(self.key.encode("utf8"), AES.MODE_CBC, iv().encode("utf8"))
        enc = cipher.encrypt(raw)
        return b64encode(enc).decode("utf8")

    # decrypt : 메시지를 복호화 하는 함수
    def decrypt(self, enc):
        enc = b64decode(enc)
        cipher = AES.new(self.key.encode("utf8"), AES.MODE_CBC, iv().encode("utf8"))
        dec = cipher.decrypt(enc)
        return unpad(dec).decode("utf8")


# 암호화 키
import random

key_id = str(random.randint(0, 2 ^ 64)).zfill(32)
key_pw = str(random.randint(0, 2 ^ 64)).zfill(32)

from selenium import webdriver


class Ui_Dialog(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        self.setupUi()

    def setupUi(self):
        self.setWindowTitle("ZEUSLoginPage")
        self.resize(400, 500)
        self.setMinimumSize(QtCore.QSize(400, 500))
        self.setMaximumSize(QtCore.QSize(400, 500))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(76, 76, 76))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(76, 76, 76))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(76, 76, 76))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(76, 76, 76))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.setPalette(palette)
        self.TitleLabel = QtWidgets.QLabel(self)
        self.TitleLabel.setGeometry(QtCore.QRect(30, 30, 340, 80))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        self.TitleLabel.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("고도 B")
        font.setPointSize(36)
        self.TitleLabel.setFont(font)
        self.TitleLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.TitleLabel.setObjectName("TitleLabel")
        self.label = QtWidgets.QLabel(self)
        self.label.setGeometry(QtCore.QRect(20, 180, 81, 41))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        self.label.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self)
        self.label_2.setGeometry(QtCore.QRect(20, 240, 81, 41))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        self.label_2.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("나눔스퀘어 Bold")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.lineEdit_ID = QtWidgets.QLineEdit(self)
        self.lineEdit_ID.setGeometry(QtCore.QRect(110, 180, 271, 41))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setPointSize(16)
        self.lineEdit_ID.setFont(font)
        self.lineEdit_ID.setObjectName("lineEdit_ID")
        self.lineEdit_PW = QtWidgets.QLineEdit(self)
        self.lineEdit_PW.setGeometry(QtCore.QRect(110, 240, 271, 41))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setPointSize(16)
        self.lineEdit_PW.setFont(font)
        self.lineEdit_PW.setEchoMode(QtWidgets.QLineEdit.Password)
        self.lineEdit_PW.setObjectName("lineEdit_PW")
        self.line = QtWidgets.QFrame(self)
        self.line.setGeometry(QtCore.QRect(20, 125, 361, 16))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.line_2 = QtWidgets.QFrame(self)
        self.line_2.setGeometry(QtCore.QRect(20, 130, 361, 16))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.pushButton = QtWidgets.QPushButton(self)
        self.pushButton.setGeometry(QtCore.QRect(20, 300, 361, 41))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setPointSize(10)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        self.label_3 = QtWidgets.QLabel(self)
        self.label_3.setGeometry(QtCore.QRect(25, 340, 351, 141))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        self.label_3.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setScaledContents(False)
        self.label_3.setWordWrap(True)
        self.label_3.setObjectName("label_3")

        self.retranslateUi(self)
        QtCore.QMetaObject.connectSlotsByName(self)
        # 여기서부터 수정한 부분
        # 타이틀바의 도움말 버튼 없애기+항상 맨 앞 창에 위치하기
        self.setWindowFlags(QtCore.Qt.WindowStaysOnTopHint)
        # 최소화 버튼 비활성화
        self.setWindowFlags(QtCore.Qt.WindowCloseButtonHint)
        # 버튼 기능 연결
        self.pushButton.clicked.connect(self.PushLoginButton)  # 다음 버튼

    def PushLoginButton(self):
        from os import getcwd
        import time
        from os.path import isfile
        program_directory = getcwd()[:getcwd().index('Design')]
        current_driver_path = program_directory + 'Design file\\chromedriver.exe'
        print(current_driver_path)
        current_driver_path = current_driver_path.replace("\\", "/")[2:]
        driver = webdriver.Chrome(current_driver_path)

        driver.get('https://zeus.gist.ac.kr/sys/main/login.do')
        time.sleep(1)

        id = self.lineEdit_ID.text()
        pw = self.lineEdit_PW.text()

        driver.find_element_by_xpath('//*[@id="login_id"]').send_keys(id)
        driver.find_element_by_xpath('//*[@id="login_pw"]').send_keys(pw)

        driver.find_element_by_xpath('//*[@id="login_wrap"]/div[2]/div[1]/fieldset/form/ul[3]/li[1]/button').click()
        time.sleep(0.5)
        no_error_token = 0  # 로그인 성공 토큰(except 발생 시 로그인 성공)
        try:
            result = driver.switch_to.alert
            result.accept()
        except:
            no_error_token = 1

        driver.close()
        global return_answer
        if no_error_token == 1:  # 로그인 성공 시
            # ID 및 비밀번호를 엑셀 파일에 저장(이후 start2.py에서 자동으로 암호화)
            from openpyxl import Workbook
            write_wb = Workbook()
            write_ws = write_wb.active
            write_ws.append(['ID', id])
            write_ws.append(['PW', pw])
            write_wb.save(program_directory + '/edited/info/personalinfo.xlsx')
            return_answer = 100
        else:  # 로그인 실패 시
            return_answer = 50
        self.close()

        if return_answer == 100:
            # ID 및 비밀번호 암호화하여 personalinfo.xlsx에 넣기
            from openpyxl import load_workbook, Workbook
            load_wb = load_workbook(program_directory + 'edited/info/personalinfo.xlsx', data_only=True)
            load_ws = load_wb['Sheet']
            write_wb = Workbook()
            write_ws = write_wb.active
            encrypted_id = AESCipher(key_id).encrypt(load_ws['B1'].value)
            encrypted_pw = AESCipher(key_pw).encrypt(load_ws['B2'].value)
            write_ws.append(['ID', encrypted_id])
            write_ws.append(['PW', encrypted_pw])
            write_wb.save(program_directory + '/edited/info/personalinfo.xlsx')
        return 0

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.TitleLabel.setText(_translate("Dialog", "ZEUS Login"))
        self.label.setText(_translate("Dialog", "ID   :"))
        self.label_2.setText(_translate("Dialog", "PW :"))
        self.pushButton.setText(_translate("Dialog", "Log-In"))
        self.label_3.setText(_translate("Dialog",
                                        "ID 및 비밀번호는 edited/info 폴더 내 personalinfo.xlsx에서 암호화된 텍스트 상태로 저장되며,"
                                        " 정상적으로 프로그램을 종료할 시 자동으로 삭제됩니다.\n"
                                        "로그인 성공 여부는 하단 상태 창에서 보실 수 있습니다."))

