def CheckNecFile(class1):
    from os import getcwd
    from os.path import isfile
    # 성적 및 교과목 요약 파일 불러오기(없으면 알림)
    # 성적 요약 파일 불러오기
    program_directory = getcwd()[:getcwd().index('Design')]
    # 교양 과목 요약 파일 불러오기
    elect_list = ('ppe', 'gsc', 'hus')
    elect_token = True
    for elect_file_name in elect_list:
        elect_edited_file = program_directory + 'edited\\elective_list\\' + elect_file_name + '.xlsx'
        if isfile(elect_edited_file) == 0:
            elect_token = False
    elect_db_file = program_directory + 'edited\\elective_list\\DBinfo.xlsx'
    if isfile(elect_db_file) == 0:
        elect_token = False
    if elect_token:
        print_text = '성적 요약 파일 및 학생 정보 파일\n(이름, 학번, 전공, 이수한 학기수)' \
                     '\nStep 1, 2 순서대로 실시(필요 시 3 눌러 교양 과목 DB 업데이트)'
        must_step_count = 2
    else:
        print_text = '성적 파일, 학생 정보 파일(이름, 학번, 전공, 이수한 학기수) 및\n교양 과목 DB' \
                     '\nStep 1, 2, 3 순서대로 실시'
        must_step_count = 3


    fig1 = isfile(program_directory + 'edited\\info\\first_graph.html')
    fig2 = isfile(program_directory + 'edited\\info\\second_graph.html')
    fig3 = isfile(program_directory + 'edited\\info\\first_table.html')
    fig4 = isfile(program_directory + 'edited\\info\\second_table.html')
    student_info = isfile(program_directory + 'edited\\info\\personalinfo2.xlsx')
    next_button_on = False
    if fig1 and fig2 and fig3 and fig4 and student_info:
        print_text = '그래프 html 파일 및 학생 정보 파일 존재함.\n' \
                     '하단의 \'다음\' 버튼 눌러 결과표 보기 가능\n' \
                     '(필요 시 Step 1, 2(+3) 순서대로 실시)'
        next_button_on = True
    return print_text, must_step_count, next_button_on


# 암호화 및 복호화 세팅
from base64 import b64encode, b64decode
from Crypto.Cipher import AES

# 임의의 값으로 패딩
BS = 16
pad = lambda s: s + (BS - len(s) % BS) * chr(BS - len(s) % BS).encode()
unpad = lambda s: s[:-ord(s[len(s) - 1:])]


def iv():
    return chr(0) * 16


class AESCipher(object):
    def __init__(self, key):
        self.key = key

    # encrypt : 메시지를 암호화 하는 함수
    def encrypt(self, message):
        message = message.encode()
        raw = pad(message)
        cipher = AES.new(self.key.encode("utf8"), AES.MODE_CBC, iv().encode("utf8"))
        enc = cipher.encrypt(raw)
        return b64encode(enc).decode("utf8")

    # decrypt : 메시지를 복호화 하는 함수
    def decrypt(self, enc):
        enc = b64decode(enc)
        cipher = AES.new(self.key.encode("utf8"), AES.MODE_CBC, iv().encode("utf8"))
        dec = cipher.decrypt(enc)
        return unpad(dec).decode("utf8")


# 암호화 키

key_id = 0
key_pw = 0


def PushStep1(class1):
    import Login2
    print('processing')

    class1.cprint("=" * 35)
    class1.cprint("1. ZEUS 로그인 가능 여부 확인 중...\n")

    loginwindow = Login2.Ui_Dialog()
    loginwindow.show()
    loginwindow.exec()
    reply_value = Login2.return_answer
    login2_key_id = Login2.key_id
    login2_key_pw = Login2.key_pw
    return reply_value, login2_key_id, login2_key_pw


def PushStep2(class1, parameter_key_id, parameter_key_pw):
    from os import getcwd, listdir, remove, rename
    from os.path import isfile, getctime
    from selenium import webdriver
    from pandas import read_excel
    import time
    import pyautogui
    import shutil

    class1.cprint("=" * 35)
    class1.cprint("2. 성적표 및 현재 수강 과목 조회 중...\n")
    
    program_directory = getcwd()[:getcwd().index('Design')]
    current_driver_path = program_directory + "\\Design File\\" + 'chromedriver'
    current_driver_path = current_driver_path.replace("\\", "/")[2:]

    # 로그인 파일 존재 여부 확인
    info_file = program_directory + 'edited/info/personalinfo.xlsx'
    if isfile(info_file) == False:
        class1.cprint("로그인 파일이 존재하지 않습니다. Step 1을 눌러 다시 로그인하십시오.")
        return 1
    driver = webdriver.Chrome(current_driver_path)
    driver.get('https://zeus.gist.ac.kr/sys/main/login.do')
    # 웹의 크기 및 위치 설정
    driver.set_window_size(1350, 800)
    driver.set_window_position(-7, -1)
    time.sleep(1)

    from openpyxl import load_workbook
    load_wb = load_workbook(program_directory + 'edited/info/personalinfo.xlsx', data_only=True)
    load_ws = load_wb['Sheet']
    decrypted_id = AESCipher(parameter_key_id).decrypt(load_ws['B1'].value)
    decrypted_pw = AESCipher(parameter_key_pw).decrypt(load_ws['B2'].value)
    driver.find_element_by_xpath('//*[@id="login_id"]').send_keys(decrypted_id)
    driver.find_element_by_xpath('//*[@id="login_pw"]').send_keys(decrypted_pw)
    driver.find_element_by_xpath('//*[@id="login_wrap"]/div[2]/div[1]/fieldset/form/ul[3]/li[1]/button').click()

    time.sleep(14)  # 웹 로딩 동안 기다리기

    pyautogui.moveTo(750, 800)  # 비밀번호 변경창 생성 문제 해결을 위한 클릭
    pyautogui.click()
    time.sleep(0.5)

    # 현재 수강하고 있는 과목 수집하여 정리
    pyautogui.moveTo(50, 330)
    pyautogui.click()
    time.sleep(0.5)
    pyautogui.moveTo(50, 440)
    pyautogui.click()
    time.sleep(3)
    pyautogui.moveTo(50, 330)
    pyautogui.click()
    time.sleep(0.5)

    # ZEUS 현재 수강과목 파싱
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    from datetime import datetime

    today = datetime.today()  # 현재 날짜 가져오기
    driver.switch_to.frame("TOBE_JSP")

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    total_num_pointer = '#mainframe_VFrameSet_HFrameSet_MDIFrameSet_ctxFrameSet_ctxFrame_' \
                        'PERS07\^PERS07_02\^003\^UlsTlsnAplyCtntQ_form_div_sample_divMain_' \
                        'divRowCnt_grxMain_stcRowCntTextBoxElement > div'
    try:
        total_num = int(str(soup.select(total_num_pointer)[0])[:-6].split('>')[-1])
    except:
        pyautogui.moveTo(50, 330)
        pyautogui.click()
        time.sleep(0.5)
        pyautogui.moveTo(50, 440 + 40)
        pyautogui.click()
        time.sleep(3)
        pyautogui.moveTo(50, 330)
        pyautogui.click()
        time.sleep(0.5)

        today = datetime.today()  # 현재 날짜 가져오기

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')

        total_num_pointer = '#mainframe_VFrameSet_HFrameSet_MDIFrameSet_ctxFrameSet_ctxFrame_' \
                            'PERS07\^PERS07_02\^003\^UlsTlsnAplyCtntQ_form_div_sample_divMain_' \
                            'divRowCnt_grxMain_stcRowCntTextBoxElement > div'

        total_num = int(str(soup.select(total_num_pointer)[0])[:-6].split('>')[-1])
        print(total_num)

    total_present_lecture = []  # 현재 수강과목 내역 리스트 수집소
    retake_lecture_code = []
    for num in range(0, total_num):
        present_lecture_string_pointer = '#mainframe_VFrameSet_HFrameSet_MDIFrameSet_ctxFrameSet_ctxFrame_' \
                                         'PERS07\^PERS07_02\^003\^UlsTlsnAplyCtntQ_form_div_sample_' \
                                         'divMain_grxMain_body_gridrow_' + str(num)
        present_lecture_string = str(soup.select(present_lecture_string_pointer)[0])

        temp_content_list = present_lecture_string.split('ellipsis;">')[1:]
        temp_content_list2 = []
        temp_content_list3 = []
        for content in temp_content_list:
            refined = content.split('</div></div></div>')[0]
            temp_content_list2.append(refined)
        temp_content_list2[0] = int(temp_content_list2[0])
        temp_content_list2[1] = temp_content_list2[1][:-3]
        # temp_content_list3에 필요한 정보만 넣기
        criteria_date = [datetime(today.year, 1, 1), datetime(today.year, 2, 10), datetime(today.year, 8, 10)]
        # if temp_content_list2[-2] == '':  # 드롭하지 않은 과목만 투입(취소: 드롭 여부 알 수 없음)
        temp_content_list3.append(temp_content_list2[1])  # 교과목(코드)
        temp_content_list3.append(int(temp_content_list2[7][-1]))  # 학점
        temp_content_list3.append(temp_content_list2[2])  # 교과목명
        if temp_content_list2[12] != "":
            retake_lecture_code.append(temp_content_list2[12][:temp_content_list2[12].find("-")])  # 재수강 이전 과목 코드
        if criteria_date[0] <= today < criteria_date[1]:
            date_content = str(today.year - 1) + '년 2학기'
        elif criteria_date[1] <= today < criteria_date[2]:
            date_content = str(today.year) + '년 1학기'
        else:
            date_content = str(today.year) + '년 2학기'
        temp_content_list3.append(date_content)
        total_present_lecture.append(temp_content_list3)

    # 재수강 과목과 그 이전과목 충돌 시 이전과목 삭제하는 코드 작성 필요(retake_lecture_code 에 존재)

    # 개인성적 엑셀 파일 다운로드 및 이름 수정하여 프로그램 폴더로 이동하기
    # '개인성적조회' 이동
    pyautogui.moveTo(50, 365)
    pyautogui.click()
    time.sleep(0.5)
    # 온라인 강의로 인한 배너 추가로 인한 y축값 변경
    pyautogui.moveTo(50, 365 + 40 * 3)
    pyautogui.click()
    # 이름 및 전공 추출
    time.sleep(10)
    from tkinter import Tk
    time.sleep(2)
    # 컴퓨터 환경에 따라 클릭이 잘 먹히지 않을 수 있음.
    for i in (1, 5):
        pyautogui.moveTo(550, 180)
        pyautogui.click()
        time.sleep(0.1)
        pyautogui.click()
        time.sleep(0.1)
        pyautogui.click()
        time.sleep(0.1)
        pyautogui.click()
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.5)
        student_name_major = Tk().clipboard_get()
        if ')' in student_name_major:
            class1.cprint("Your name and major have been extracted.")
            break
        else:
            class1.cprint("Extracting your name and major(%d/5)..." % i)
        if i == 5:
            return 1
    student_name = student_name_major[:student_name_major.find('(')]  # 학생 이름
    student_major = student_name_major[student_name_major.find('(') + 1:student_name_major.find('-')]  # 학생 전공
    if student_major == "원외접속":
        student_major = student_name_major[student_name_major.find('-') + 1:student_name_major.rfind('-')]  # 학생 전공
    student_status = student_name_major[student_name_major.find('-') + 1:student_name_major.find(')')]
    # 학생 상태(재학, 휴학, 졸업 등)
    # report card 창 생성
    time.sleep(1)
    pyautogui.moveTo(1300, 515)
    pyautogui.click()
    # 엑셀 파일 다운로드
    time.sleep(3)
    pyautogui.moveTo(355, 320)
    pyautogui.click()
    time.sleep(5)

    # 엑셀 파일 옮기기 및 이름 바꾸기
    import getpass
    username = getpass.getuser()
    print("사용 계정: "+username)
    grade_path = 'C:/Users/'+username+'/Downloads'
    grade_path_list = listdir(grade_path)
    grade_filename = list(filter(lambda x: x.startswith('Completed course grade'), grade_path_list))
    if len(grade_filename) == 0:
        return 2

    # 가장 최신 성적 파일 선정 및 파일을 프로그램 관련 폴더로 이동한 후 파일명 수정하기
    recent_filename = 'Completed course grade.xls'
    for file in grade_filename:
        grade_file = grade_path + '/' + file
        pre_grade_file = grade_path + '/' + recent_filename
        final_access_time = getctime(grade_file)
        if final_access_time >= getctime(pre_grade_file):
            recent_filename = file

    class1.cprint('\nDownloading file: ' + recent_filename)

    recent_file_path = grade_path + '/' + recent_filename
    db_file_directory = program_directory.replace("\\", "/")[2:] + '/original/grade'

    if isfile(db_file_directory + '/grade.xls'):
        remove(db_file_directory + '/grade.xls')

    shutil.move(recent_file_path, db_file_directory)
    rename(db_file_directory + '/' + recent_filename, db_file_directory + '/' + 'grade.xls')
    class1.cprint('Renaming file: ' + recent_filename + ' => ' + 'grade.xls')
    driver.close()

    # 다운받은 성적표를 정리
    grade_file = program_directory + 'original\\grade\\grade.xls'
    df_grade = read_excel(grade_file)

    # 학번 추출
    student_no = int(df_grade.loc[[0], 'Unnamed: 0'].apply(lambda x: x.split(' : ')[-1].strip()))

    # 성적표 엑셀 파일 수정
    df_grade = df_grade.loc[:, ['Unnamed: 1', 'Unnamed: 3', 'Unnamed: 4', 'Unnamed: 5']]

    # 엑셀 파일에 수강 기간 열 추가
    grade_row_count = len(df_grade)

    course_date_list = []
    semester_count_list = []
    year_semester = "asdf"
    semester_count = 0
    exist_ap = False

    for row in range(0, grade_row_count):
        semester_cell = str(df_grade.loc[row, 'Unnamed: 3']).strip()
        if semester_cell[0] == '<' and semester_cell[-1] == '>':
            if semester_cell.find("1학기") == -1 and semester_cell.find("2학기") == -1:  # 인정학기(AP), 여름학기 또는 겨울학기 수강 시
                if semester_cell.find("인정") != -1:  # 인정학기일 시
                    year_semester = semester_cell[1:5] + '년 1학기'
                    semester_count += 1
                    exist_ap = True
                else:  # 계절학기일 시
                    year_semester = semester_cell[1:5] + '년 ' + course_date_list[-1][-3:]
            else:  # 정규학기 수강 시
                year_semester = semester_cell[1:5] + '년 ' + semester_cell[6:9]
                semester_count += 1
                if exist_ap:
                    exist_ap = False
                    semester_count -= 1

        course_date_list.append(year_semester)
        semester_count_list.append(semester_count)

    df_grade['수강 학기'] = course_date_list
    df_grade['수강 학기수'] = semester_count_list

    # df_grade_for_save = df_grade.dropna(axis=0)[1:].reset_index(drop=True)
    df_grade_for_save = df_grade.reset_index(drop=True)
    # 행 이름 변경
    df_grade_for_save.columns = ['교과목', '교과목명', '학점', '평점', '수강 학기', '수강 학기수']

    # 행 순서 변경(시간표 목록 순서대로)
    df_grade_for_save = df_grade_for_save[['교과목', '학점', '교과목명', '수강 학기', '수강 학기수', '평점']]

    # 아직 성적 반영되지 않은 과목 S로 처리하기
    df_grade_for_save['평점'] = df_grade_for_save['평점'].fillna('S')
    df_grade_for_save = df_grade_for_save.dropna(axis=0)[1:].reset_index(drop=True)

    # 학점을 string형에서 integer형으로 변경
    df_grade_for_save['학점'] = df_grade_for_save['학점'].apply(lambda x: int(x))

    df_grade_for_save.to_excel(program_directory + 'edited\\grade\\grade_edited.xlsx', index=False)

    # 이전에 수강한 과목을 현재 수강하고 있는 과목과 통합하여 저장
    df_grade_for_save2 = df_grade_for_save[['교과목', '학점', '교과목명', '수강 학기']]
    for array in total_present_lecture:
        df_grade_for_save2.loc[len(df_grade_for_save2)] = array
    print(df_grade_for_save2)

    df_grade_for_save2.to_excel(program_directory + 'edited\\grade\\present_subject_info.xlsx', index=False)

    # personalinfo2.xlsx에 학생 정보 추가
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws.append(['Student No', student_no])
    write_ws.append(['Semester(s)', semester_count])
    write_ws.append(['Name', student_name])
    write_ws.append(['Major', student_major])
    write_wb.save(program_directory + '/edited/info/personalinfo2.xlsx')

    class1.cprint("\ngrade_edited.xlsx saved in " + program_directory + 'edited\\grade folder.')
    class1.cprint('\nSummary of grade file has done.')
    class1.cprint('성적표 및 학생 정보 다운로드를 완료했습니다. 다음 단계를 진행하십시오.')
    return 0


def PushStep3(class1):
    import os
    from selenium import webdriver
    import time
    import shutil
    class1.cprint("=" * 35)
    class1.cprint("3. 교양 과목 데이터베이스 업데이트 중...\n")
    program_directory = os.getcwd()[:os.getcwd().index('Design')]
    current_driver_path = program_directory + "\\Design File\\" + 'chromedriver'
    current_driver_path = current_driver_path.replace("\\", "/")[2:]
    driver = webdriver.Chrome(current_driver_path)
    elective_list = ['gsc', 'ppe', 'hus']
    for elective in elective_list:
        driver.get('https://github.com/AlpacaParker4592/GIST_Credit_Analysis_Program_without_IDE/blob/master/edited'
                   '/elective_list/'+elective+'.xlsx')
        time.sleep(3)
        driver.find_element_by_xpath('//*[@id="raw-url"]').click()
        time.sleep(2)
        # 엑셀 파일 옮기기 및 이름 바꾸기
        import getpass
        username = getpass.getuser()
        print("사용 계정: " + username)
        grade_path = 'C:/Users/' + username + '/Downloads'
        grade_path_list = os.listdir(grade_path)
        grade_filename = list(filter(lambda x: x.startswith(elective), grade_path_list))
        for i in range(0, 6):
            grade_filename = list(filter(lambda x: x.startswith(elective), grade_path_list))
            if len(grade_filename) == 0:
                if i == 5:
                    return 1
                class1.cprint("다운로드 재확인 중...(" + str(i + 1) + "/5)")
                time.sleep(1)
            else:
                break

        # 가장 최신 DB 파일 선정 및 파일을 프로그램 관련 폴더로 이동한 후 파일명 수정하기
        recent_filename = elective+'.xlsx'
        for file in grade_filename:
            grade_file = grade_path + '/' + file
            pre_grade_file = grade_path + '/' + recent_filename
            final_access_time = os.path.getctime(grade_file)
            if final_access_time >= os.path.getctime(pre_grade_file):
                recent_filename = file

        class1.cprint('Downloading file: ' + recent_filename)

        recent_file_path = grade_path + '/' + recent_filename
        db_file_directory = program_directory.replace("\\", "/")[2:] + '/edited/elective_list'

        if os.path.isfile(db_file_directory + '/'+elective+'.xlsx'):
            os.remove(db_file_directory + '/'+elective+'.xlsx')

        shutil.move(recent_file_path, db_file_directory)
        os.rename(db_file_directory + '/' + recent_filename, db_file_directory + '/' + elective + '.xlsx')
        class1.cprint('Renaming file: ' + recent_filename + ' => ' + elective + '.xlsx')
    driver.close()
    
    # DB info 엑셀파일 제작
    from datetime import datetime
    from openpyxl import Workbook, load_workbook
    updated_date = datetime.today().strftime("%Y.%m.%d")  # YYYY/mm/dd 형태의 시간 출력
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws.append(['DB Updated Date', updated_date])
    # 교양 hus, ppe 적용일자 입력
    load_wb = load_workbook(program_directory + 'edited/elective_list/hus.xlsx', data_only=True)
    load_ws = load_wb['Sheet1']
    DB_applied_year = load_ws['C2'].value
    write_ws.append(['HUS, PPE Applied Year', DB_applied_year])
    # DB 저장
    write_wb.save(program_directory + '/edited/elective_list/DBinfo.xlsx')
    class1.cprint("\nDBinfo.xlsx saved in " + 'edited\\elective_list folder.')
    class1.cprint('\n교양 과목 DB 업데이트를 완료했습니다. 하단의 다음 버튼을 누르십시오.')
    class1.cprint("최신 DB 업데이트 날짜: " + updated_date)
    class1.cprint("교양 과목 DB 적용 연도: 2014 ~ " + str(DB_applied_year)+"년")
    return 0

